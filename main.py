from fastapi import FastAPI, Request, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from fastapi.responses import FileResponse
import time
from Jobs import Jobs
import os
from fastapi.staticfiles import StaticFiles

# Initialize the FastAPI app
app = FastAPI()

# Set up Jinja2 templates directory
templates = Jinja2Templates(directory="templates")

# Mount static files directory for serving static content
app.mount("/static", StaticFiles(directory="static"), name="static")

# Dictionary to keep track of the processing status of each job
processing_status = dict()

# Function to delete a file after a delay


def delete_file(file_path):
    time.sleep(5)
    if os.path.exists(file_path):
        global processing_status
        del processing_status[file_path]
        os.remove(file_path)

# Function to delete a file after a delay


def delete_file2(file_path):
    time.sleep(3600)
    if os.path.exists(file_path):
        time.sleep(4)
        global processing_status
        del processing_status[file_path]
        os.remove(file_path)

# Asynchronous function to process the job


async def process_job(name, job_title: str, location: str, time2: str, count: int):
    global processing_status
    processing_status[name] = 'pending'
    j = Jobs(name, job_title, location, time2, count)
    j.found_all_works()
    time.sleep(1)
    processing_status[name] = "completed"

# Endpoint for the main page


@app.get("/", response_class=HTMLResponse)
async def main(request: Request):
    return templates.TemplateResponse("main.html", {"request": request})

# Endpoint to submit a new job


@app.post("/submit-job")
async def submit_job(request: Request, background_tasks: BackgroundTasks):
    name = str(request.client.host) + "-" + str(request.client.port)+".xlsx"
    form_data = await request.form()
    job_title = form_data.get("job_title")
    location = form_data.get("location")
    time_value = form_data.get("time")
    count = form_data.get("count")

    # Start job processing in the background
    background_tasks.add_task(
        process_job, name, job_title, location, time_value, count)

    # Redirect user to the waiting page
    return templates.TemplateResponse("wait.html", {"request": request, 'name': name})

# Endpoint to display the results of the job


@app.get("/results/{name}", response_class=HTMLResponse)
async def results(request: Request, name, background_tasks: BackgroundTasks):
    excel_file_path = name
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active
    background_tasks.add_task(delete_file2, excel_file_path)
    return templates.TemplateResponse("result.html", {"request": request, 'name': name, "exel": list(sheet.iter_rows(values_only=True))[1:]})

# Endpoint to download the Excel file


@app.get("/download-file/{name}")
async def download_file(name, background_tasks: BackgroundTasks):
    file_path = name
    file_name = "Jobs.xlsx"
    background_tasks.add_task(delete_file, file_path)
    return FileResponse(file_path, filename=file_name)

# Endpoint to check the processing status of the job


@app.get("/check-status/{name}")
async def check_status(name):
    # Return the current processing status
    return JSONResponse({"status": processing_status[name]})
