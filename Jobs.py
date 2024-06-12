from selenium import webdriver
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side
from selenium.webdriver.chrome.options import Options
import time
import sys

# Importing job search functions from different modules
from linkedin import linkedin_info
from indeed import indeed_info
from stepstone import stepstone_info

# Define a class to manage job search information


class Jobs:
    def __init__(self, name, job_title, loc, time, count):
        self.name = name  # Name of the Excel file
        self.job_title = job_title  # Job title to search for
        self.loc = loc  # Location to search for jobs
        self.time = time  # Time range for job search
        self.count = count  # Maximum number of jobs to find

    # Method to create a new Excel file with a header row
    def create_exel(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Jobs_info'  # Set the title of the sheet
        sheet.append(['Job title', 'Company name', 'Location',
                     'Description', "Link"])  # Add header row

        header_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Set header fill color
        for cell in sheet[1]:  # Apply the fill color to each cell in the header row
            cell.fill = header_fill
        wb.save(self.name)  # Save the workbook with the specified name

    # Method to search for jobs on Indeed and save to Excel
    def found_indeed(self, driver):
        try:
            indeed_info(self.job_title, self.loc,
                        self.name, self.time, self.count, driver)
        except Exception as e:
            print(f"Error in Indeed: {e}")

    # Method to search for jobs on Stepstone and save to Excel
    def found_stepstone(self, driver):
        try:
            stepstone_info(self.job_title, self.loc,
                           self.name, self.time, self.count, driver)
        except Exception as e:
            print(f"Error in Stepstone: {e}")

    # Method to search for jobs on LinkedIn and save to Excel
    def found_linkedin(self, driver):
        try:
            linkedin_info(self.job_title, self.loc,
                          self.name, self.time, self.count, driver)
        except Exception as e:
            print(f"Error in LinkedIn: {e}")

    # Method to search for jobs on all platforms and save to Excel
    def found_all_works(self):
        self.create_exel()  # Create a new Excel file
        try:
            options = Options()
            # Set browser language to English
            options.add_argument("--lang=en")
            # Set browser window size
            options.add_argument("--window-size=1920,1080")
            # Run browser in headless mode (no GUI)
            options.add_argument("--headless")
            options.add_argument(
                "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            )  # Set user agent to mimic a real browser
            driver = webdriver.Chrome(options=options)
        except:
            print("Some error")
            return 0
        self.found_indeed(driver)  # Search for jobs on Indeed
        self.found_stepstone(driver)  # Search for jobs on Stepstone
        self.found_linkedin(driver)  # Search for jobs on LinkedIn
        driver.quit()  # Quit the browser
